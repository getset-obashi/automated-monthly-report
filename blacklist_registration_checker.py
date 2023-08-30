import re
import sys
import os
import csv
import openpyxl
from openpyxl.workbook import Workbook
from datetime import datetime
from ipaddress import IPv4Address

from utils import DOWNLOAD_PATH, get_nth_weekday_of_month, get_sys_arg

# デフォルトのファイルパスとAdmin/Webの対象番号リスト
DEFAULT_FILE_PATH = os.path.join(DOWNLOAD_PATH,f"エラー報告レポート（{get_nth_weekday_of_month(0, 4, 2).strftime('%Y_%-m_%-d')}）.xlsx")
DEFAULT_ADMIN_NO_LIST = '6,7,8,9'
DEFAULT_WEB_NO_LIST = '8,9'
DEFAULT_WHITE_IP_FILE_PATH = os.path.join(DOWNLOAD_PATH,'htaccess.txt')
DEFAULT_BLACK_IP_FILE_PATH = os.path.join(DOWNLOAD_PATH,'ブラックリスト.txt')

# IP正規表現パターン
IP_ADDRESS_PATTERN = r'[0-9]+(?:\.[0-9]+){3}'

# マジックナンバー
ADMIN_SIDE_ERROR_MARKER = 'admin側エラー'
WEB_SIDE_ERROR_MARKER = 'web側エラー'

# IP振り分け結果CSVのファイル名
OUTPUT_CSV_FILENAME = f'エラーレポートIPチェック_{datetime.now():%Y%m%d}.csv'

# シート名
RESULT_SHEET_NAME = '集計結果'

# IPアドレスが有効かチェックする関数
def is_valid_ip(ip: str) -> bool:
    return ip is not None and IPv4Address(ip)

# ファイルの存在チェックを行う関数
def file_exists(item_name: str, path: str) -> None:
    if not os.path.isfile(path):
        print(f'{os.path.abspath(path)}: {item_name}は存在しません。引数を確認してください。')
        sys.exit()

# IPアドレスを読み込む関数
def load_ip_addresses(file_path: str) -> list:
    with open(file_path, 'r') as file:
        data = file.read()
    return find_ip_addresses(data)

# 文字列からIPアドレスを抽出する関数
def find_ip_addresses(data: str) -> list:
    ip_list = re.findall(IP_ADDRESS_PATTERN, data)
    return [ip for ip in ip_list if is_valid_ip(ip)]

# 文字列を整数に変換する関数
def try_parse_int(s: str) -> tuple:
    if s is None:
        return False, 0
    try:
        parsed_value = int(s)
        return True, parsed_value
    except ValueError:
        return False, 0

# ワークブック内に指定されたシートが存在するかチェックする関数
def any_worksheet_exists(workbook: Workbook, sheet_name: str) -> bool:
    return any(ws.title == sheet_name for ws in workbook.worksheets)

# リストの要素を一行ずつ出力する関数
def print_list(items: list):
    for item in items:
        print(item)

# リスト内の文字列のいずれかが指定された文字列を含んでいるかチェックする関数
def is_string_list_any(str_list: list, key: str) -> bool:
    return any(key in item for item in str_list)

# ホワイトリストを読み込み、リストとして返す関数
def load_white_list(file_path: str) -> list:
    print(f'ホワイトリスト読み込み path={os.path.abspath(file_path)}')
    ip_white_list = load_ip_addresses(file_path)
    print(f'ホワイトリスト読み込み結果')
    print_list(ip_white_list)
    return ip_white_list

# ブラックリストを読み込み、リストとして返す関数
def load_black_list(file_path: str) -> list:
    print(f'ブラックリスト読み込み path={os.path.abspath(file_path)}')
    ip_black_list = load_ip_addresses(file_path)
    print(f'ブラックリスト読み込み結果')
    print_list(ip_black_list)
    return ip_black_list

# エラー報告レポートから該当エラーのIPアドレスを抽出し、リストとして返す関数
def extract_error_ips(file_path: str, admin_no_list: list, web_no_list:list) -> tuple:
    print(f'該当エラーIP抽出 path={os.path.abspath(file_path)}')
    wb = openpyxl.load_workbook(file_path)

    admin_target_list = {}
    web_target_list = {}
    target_sheet_name = RESULT_SHEET_NAME
    if not any_worksheet_exists(wb, target_sheet_name):
        print(f'指定されたシート "{target_sheet_name}" は存在しません。')
        sys.exit()

    ws = wb[target_sheet_name]
    admin_flg = False
    web_flg = False

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == ADMIN_SIDE_ERROR_MARKER:
            admin_flg = True
            continue
        if row[0] == WEB_SIDE_ERROR_MARKER:
            admin_flg = False
            web_flg = True
            continue

        result, no = try_parse_int(row[1])
        if result:
            def execute(target_list, res_dic):
                if no in target_list:
                    items = []
                    for item in row[9:]:
                        if item is None:
                            break
                        items.append(item.replace('*', '.*'))
                    res_dic[no] = items

            if admin_flg:
                execute(admin_no_list, admin_target_list)
            if web_flg:
                execute(web_no_list, web_target_list)

    date_str = re.findall('\（.+?\）', os.path.basename(file_path))[0]
    target_sheet_name = datetime.strptime(date_str, "（%Y_%m_%d）").strftime('%Y%m%d')
    
    if any_worksheet_exists(wb, target_sheet_name):
        ws = wb[target_sheet_name]
        admin_res_dic = {}
        web_res_dic = {}
        ip_list = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue

            cell = row[6]
            server = row[4]
            value = cell
            if value is None:
                continue

            def execute(target_list, res_dic):
                for no in target_list:
                    for target in target_list[no]:
                        if re.match(target, value):
                            if not no in res_dic:
                                res_dic[no] = []
                            print(f'{server} No.{no} {value}')
                            res_dic[no].append(value)
                            ip_addresses = find_ip_addresses(value)
                            if len(ip_addresses) > 0:
                                if ip_addresses[0] == '127.0.0.1':
                                    if len(ip_addresses) > 1:
                                        ip_list.append(ip_addresses[1])
                                    else:
                                        print(f'対象のエラーログからIPアドレスを抽出できませんでした:{server} No.{no} {value}')
                                else:
                                    ip_list.append(ip_addresses[0])
                            else:
                                print(f'対象のエラーログからIPアドレスを抽出できませんでした:{server} No.{no} {value}')

            if server == 'admin':
                execute(admin_target_list, admin_res_dic)
            if server == 'web':
                execute(web_target_list, web_res_dic)
        wb.close()
        
        ip_list = sorted(set(ip_list), key=IPv4Address)
        print(f'該当エラーIP抽出結果')
        print_list(ip_list)
        return ip_list,admin_res_dic,web_res_dic
    else:
        print(f'指定されたシート "{target_sheet_name}" は存在しません。')
        return [],{},{}

# エラーのIPアドレスをカテゴリー分けし、リストとして返す関数
def categorize_ips(ip_list: list, ip_white_list: list, ip_black_list: list, admin_res_dic: dict, web_res_dic: dict) -> tuple:
    res_white_ip_list = []
    res_black_ip_list = []
    res_black_ip_list_regist_request = []

    print(f'該当エラーIP振り分け')
    for ip in ip_list:
        if ip in ip_white_list:
            res_white_ip_list.append(ip)
            message = 'ホワイトリストに含まれているため追加不要'
        elif ip in ip_black_list:
            res_black_ip_list.append(ip)
            message = '既にブラックリストに含まれているため追加不要'
        else:
            res_black_ip_list_regist_request.append(ip)
            message = 'ブラックリスト登録依頼'
        
        div = []
        for no in admin_res_dic:
            if is_string_list_any(admin_res_dic[no], ip):
                div.append(f'Admin No.{no}')
        for no in web_res_dic:
            if is_string_list_any(web_res_dic[no], ip):
                div.append(f'Web No.{no}')
        print(f'{message}({",".join(div)}): {ip}')

    return res_white_ip_list, res_black_ip_list, res_black_ip_list_regist_request

# 結果をCSVファイルに書き込む関数
def write_to_csv(res_white_ip_list: list, res_black_ip_list: list, res_black_ip_list_regist_request: list):
    with open(OUTPUT_CSV_FILENAME, 'w+', newline='', encoding='cp932') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['ホワイトリストに含まれているIP', '既にブラックリストに含まれているIP', 'ブラックリスト登録依頼IP'])
        count_list = []
        for idx in range(max([len(res_white_ip_list), len(res_black_ip_list), len(res_black_ip_list_regist_request)])):
            def get_item(item_list: list):
                if len(item_list) > idx:
                    return item_list[idx]
                return ''
            writer.writerow([get_item(res_white_ip_list), get_item(res_black_ip_list), get_item(res_black_ip_list_regist_request)])

# メイン関数
def main():
    # ファイルパスと管理者/ウェブの番号リストを取得
    file_path = get_sys_arg(1, DEFAULT_FILE_PATH)
    admin_no_list = [int(x) for x in get_sys_arg(2, DEFAULT_ADMIN_NO_LIST).split(',')]
    web_no_list = [int(x) for x in get_sys_arg(3, DEFAULT_WEB_NO_LIST).split(',')]
    white_ip_file_path = get_sys_arg(4, DEFAULT_WHITE_IP_FILE_PATH)
    black_ip_file_path = get_sys_arg(5, DEFAULT_BLACK_IP_FILE_PATH)

    file_exists('エラー報告レポート', file_path)
    file_exists('IPホワイトリスト', white_ip_file_path)
    file_exists('IPブラックリスト', black_ip_file_path)

    ip_white_list = load_white_list(white_ip_file_path)
    ip_black_list = load_black_list(black_ip_file_path)

    ip_list, admin_res_dic, web_res_dic = extract_error_ips(file_path, admin_no_list, web_no_list)
    res_white_ip_list, res_black_ip_list, res_black_ip_list_regist_request = categorize_ips(ip_list, ip_white_list, ip_black_list, admin_res_dic, web_res_dic)
    write_to_csv(res_white_ip_list, res_black_ip_list, res_black_ip_list_regist_request)

if __name__ == "__main__":
    main()