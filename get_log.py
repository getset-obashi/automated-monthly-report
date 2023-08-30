import html
import os
from typing import List
import get_log_php
import get_log_laravel
from log_entry import LogEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime
import openpyxl.utils.exceptions as openpyxl_exceptions
from utils import get_aggregation_period, get_last_weekday_of_month, DOWNLOAD_PATH, get_nth_weekday_of_month

def sanitize_string(input_string):
    return html.escape(input_string)

def write_to_excel(entries: List[LogEntry], file_name: str, tab_name: str):
    """
    ログエントリをエクセルファイルに書き込む
    :param entries: ログエントリのリスト
    :param file_name: 出力するエクセルファイル名
    :param tab_name: 出力するタブ名
    """
    workbook = load_workbook(filename=file_name)
    sheet = workbook[tab_name]

    # Find the first empty cell in column G
    g_column = sheet['G']
    empty_row = 2
    while g_column[empty_row - 1].value is not None:
        empty_row += 1

    start_date, end_date = get_aggregation_period()
    sheet.cell(row=1, column=1, value=f'集計期間:{start_date}~{end_date}')

    # Set the starting row for data output
    starting_row = max(empty_row, 3)

    # Write the data again starting from the first empty row
    for entry in entries:
        row = [entry.date, entry.server, entry.location, entry.content]
        for idx, value in enumerate(row):
          try:
              sheet.cell(row=starting_row, column=4 + idx, value=value)
          except openpyxl_exceptions.IllegalCharacterError:
              sheet.cell(row=starting_row, column=4 + idx, value="Invalid Character")
              print(f'openpyxl_exceptions.IllegalCharacterError row:{starting_row} col:{4 + idx}\n{value}')
        starting_row += 1

    workbook.save(file_name)


def main():
    file_date: datetime = get_nth_weekday_of_month(0, 4, 2)
    formatted_file_date: str = file_date.strftime('%Y_%-m_%-d')  # フォーマットした文字列を取得
    output_file_name: str = os.path.join(DOWNLOAD_PATH,f"エラー報告レポート（{formatted_file_date}）.xlsx")
    if not os.path.exists(output_file_name):
        print(f'出力先ファイルが見つかりません。path={output_file_name}')
        return
    output_directory:str = os.path.join(os.getcwd(),datetime.today().strftime('%Y%m%d%H%M%S'))
    start_date, end_date = get_aggregation_period()
    log_entries =  get_log_php.main(output_directory, start_date, end_date)
    log_entries.extend(get_log_laravel.main(output_directory, start_date, end_date))
    tab_name: str = file_date.strftime('%Y%m%d')
    write_to_excel(log_entries, output_file_name, tab_name)


if __name__ == "__main__":
    main()